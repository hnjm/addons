from odoo import api, models, fields, _
from odoo.exceptions import UserError, ValidationError
from odoo.tools.float_utils import float_compare, float_is_zero, float_round
from collections import defaultdict
import io
import json
import re
import base64
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Color, Font, Border, Side, Alignment, fills, borders
from odoo.modules.module import get_module_resource


# as per recommendation from @freylis, compile once only
CLEANR = re.compile('<.*?>')


def cleanhtml(raw_html):
    cleantext = re.sub(CLEANR, '', raw_html)
    return cleantext


class SaleOrderReport(models.TransientModel):
    _name = "sale.order.report"

    excel_file = fields.Binary('Report file ')
    file_name = fields.Char('Excel file', size=64)

class SaleOrder(models.Model):
    _inherit = 'sale.order'

    def total_sale_order(self):
        total = json.loads(self.tax_totals_json)
        if total:
            return total['amount_total']

    def get_employee_salesperson(self):
        partner = self.env['res.users'].browse(self.user_id.id).partner_id
        return partner.name

    def get_current_week_of_year(self):
        date = datetime.now()
        return date.isocalendar()[1]

    def format_date(self, date):
        if date:
            res = str(date).split('-')
            return '{}-{}-{}'.format(res[2], res[1], res[0])

    def format_currency(self, number):
        currency = "${:,.2f}".format(number)
        return currency

    def action_generate_xlsx_quotation(self):
        filename = f'sale_quotation_confirmation_template_{self.name}'
        existing_filepath = get_module_resource('report_quotation_template', 'static',
                                                'sale_quotation_confirmation_template.xlsx')
        wb = load_workbook(existing_filepath)
        ws = wb.active

        quotation_number = f': {self.name}'
        quotation_date = f': {self.date_order.strftime("%d-%m-%Y") if self.date_order else ""}'
        quotation_validity = f': {self.validity_date.strftime("%d-%m-%Y") if self.validity_date else ""}'
        ref = f': {self.client_order_ref or ""}'
        phone = f': {self.partner_id.phone if self.partner_id else ""}'
        email = f': {self.partner_id.email if self.partner_id else ""}'
        incoterm_name = f': {self.incoterm.name or "" if self.incoterm else ""}'
        payment_term = f': {self.payment_term_id.name if self.payment_term_id else ""}'
        rev = f': {self.rev or ""}'
        week = f': {self.get_current_week_of_year() or ""}'

        # address
        address_list = []
        if self.partner_id and self.partner_id.parent_id:
            address_list.append(self.partner_id.parent_id.name)
        if self.partner_id:
            address_list.append(self.partner_id.name)
        if self.partner_id and self.partner_id.street:
            address_list.append(self.partner_id.street)

        street2 = self.partner_id and self.partner_id.street2 or ""
        city = self.partner_id.city or ""
        state_id = self.partner_id.state_id and self.partner_id.state_id.name or ""
        zip = self.partner_id.zip or ""
        street2_lst = []
        street2_lst += [street2] if street2 else []
        street2_lst += [city] if city else []
        street2_lst += [state_id] if state_id else []
        street2_lst += [zip] if zip else []
        street_str = ", ".join(street2_lst)
        address_list.append(street_str)
        if self.partner_id and self.partner_id.country_id:
            address_list.append(self.partner_id.country_id.name)
        address_str = "\n".join(address_list)

        cell = ws.cell(row=8, column=5)  # F15
        cell.value = address_str
        ws.merge_cells(start_row=8, end_row=11, start_column=5, end_column=8)


        cell = ws.cell(row=15, column=6)  # F15
        cell.value = quotation_number

        cell = ws.cell(row=16, column=6)  # F16
        cell.value = quotation_date

        cell = ws.cell(row=15, column=9)  # H15
        cell.value = rev

        cell = ws.cell(row=16, column=9)  # H16
        cell.value = week

        cell = ws.cell(row=18, column=2)  # B18
        cell.value = ref

        cell = ws.cell(row=19, column=2)  # B19
        cell.value = phone

        cell = ws.cell(row=20, column=2)  # B20
        cell.value = email

        cell = ws.cell(row=18, column=6)  # F18
        cell.value = quotation_validity

        cell = ws.cell(row=19, column=6)  # F19
        cell.value = incoterm_name

        cell = ws.cell(row=20, column=6)  # F20
        cell.value = payment_term

        # write order line
        line_orange = Border(
            top=Side(border_style=borders.BORDER_THIN, color='e99d4e'),
        )
        line_orange_bot = Border(
            bottom=Side(border_style=borders.BORDER_THIN, color='e99d4e'),
        )
        line_orange_top_bot = Border(
            top=Side(border_style=borders.BORDER_THIN, color='e99d4e'),
            bottom=Side(border_style=borders.BORDER_THIN, color='e99d4e')
        )
        num_row = 23
        index_row = 1
        total_order = 0
        for line_ in self.order_line:
            line = [
                index_row*10,
                line_.product_id.get_item_nr_report() or "",
                line_.product_id.product_tmpl_id.revision or "",
                line_.product_id.get_name_report() or "",
                "",
                "",
                line_.product_uom_qty or "" ,
                line_.price_unit or "",
                line_.price_subtotal
            ]
            index_column = 1
            for value in line:
                cell = ws.cell(row=num_row, column=index_column)
                cell.value = value
                index_column += 1
            ws.merge_cells(start_row=num_row, end_row=num_row, start_column=4, end_column=6)
            num_row += 1
            index_row += 1
            total_order += line_.price_subtotal or 0

        # write border-top 1px orange
        for i in range(1, 10):
            cell = ws.cell(row=num_row-1, column=i)
            cell.border = line_orange_bot
            cell.font = Font(size=12, name='Arial')

        # amount untaxed
        cell = ws.cell(row=num_row, column=8)  #
        cell.border = line_orange_bot
        cell.value = 'Untaxed'
        cell.alignment = Alignment(horizontal='right')
        cell.font = Font(size=12, name='Arial')

        cell = ws.cell(row=num_row, column=9)  #
        cell.border = line_orange_bot
        cell.value = total_order
        cell.font = Font(size=12, name='Arial')

        # amount taxed
        cell = ws.cell(row=num_row+1, column=8)  #
        cell.value = 'Taxed'
        cell.alignment = Alignment(horizontal='right')
        cell.font = Font(size=12, name='Arial')
        cell = ws.cell(row=num_row+1, column=9)
        cell.value = total_order + self.amount_tax

        # note
        cell = ws.cell(row=num_row+2, column=1)  #
        cell.value = cleanhtml(self.note or '')
        ws.merge_cells(start_row=num_row+2, end_row=num_row+2, start_column=1, end_column=9)
        # write border-top 1px orange
        for i in range(1, 10):
            cell = ws.cell(row=num_row+2, column=i)
            cell.border = line_orange_top_bot
            cell.font = Font(size=12, name='Arial')

        ws.merge_cells(start_row=num_row+4, end_row=num_row+4, start_column=1, end_column=5)
        cell = ws.cell(row=num_row+4, column=1)  #

        cell.value = 'We thank you for your interest and are looking forward to receive this order.'
        cell = ws.cell(row=num_row+6, column=1)  #
        cell.value = f'{self.user_id.name if self.user_id else ""}'

        cell = ws.cell(row=num_row+8, column=1)  #
        cell.value = f'{self.user_id.partner_id.function  or "" if self.user_id and self.user_id.partner_id else ""}'

        # write border-top 1px orange
        for i in range(1, 10):
            cell = ws.cell(row=num_row+16, column=i)
            cell.border = line_orange
            cell.font = Font(size=12, name='Arial')

        ws.merge_cells(start_row=num_row+16, end_row=num_row+16, start_column=1, end_column=9)
        cell = ws.cell(row=num_row+16, column=1)  #
        cell.value = u'Indochina Park tower, 19th floor, 4 Nguyen Dinh Chieu, Da Kao Ward, District 1, Ho Chi Minh, Vietnam • T: +84 28 627 55 884 • sales@abc-vietnam.com • www.abc-vietnam.com'

        ws.merge_cells(start_row=num_row+17, end_row=num_row+17, start_column=2, end_column=8)
        cell = ws.cell(row=num_row+17, column=2)  #
        cell.value = u'Factory address: 621 Hanoi highway, Group KP5, Linh Trung Ward, Thu Duc District, Ho Chi Minh, Vietnam • T: +84 28 627 55 884'

        ws.merge_cells(start_row=num_row+18, end_row=num_row+18, start_column=1, end_column=9)
        cell = ws.cell(row=num_row+18, column=1)  #
        cell.value = u'Techcombank • Tran Nao Branch • 125 Tran Nao Street, Binh An Ward, District 2, Ho Chi Minh, Vietnam • SWIFT: VTCBVNVX • VND: 19135665385012 • USD: 19135665385020'
        # write border-bottom 1px orange
        for i in range(1, 10):
            cell = ws.cell(row=num_row+18, column=i)
            cell.border = line_orange_bot
            cell.font = Font(size=12, name='Arial')

        fp = io.BytesIO()
        wb.save(fp)
        try:
            from base64 import encodebytes
        except ImportError:
            from base64 import encodestring as encodebytes

        export_id = self.env['sale.order.report'].sudo().create(
            {'excel_file': encodebytes(fp.getvalue()), 'file_name': filename})
        res = {
            'type': 'ir.actions.act_url',
            'name': filename,
            'url': '/web/content/sale.order.report/%s/excel_file/%s?download=true' % (
                export_id.id, filename),
            'target': 'new',
        }
        return res


class SaleOderLine(models.Model):
    _inherit = 'sale.order.line'

    def format_currency(self, number):
        currency = "${:,.2f}".format(number)
        return currency

    # def validity_quotation(self):
    #     if self.validity_date and self.date_order:
    #         validity_date = datetime.strptime(str(self.validity_date), '%Y-%m-%d')
    #         date_order = datetime.strptime(str(self.date_order), '%Y-%m-%d %H:%M:%S')
    #         res = validity_date.day - date_order.day
    #         return res



